import openpyxl.workbook
import requests
import json


#Configuração de excel
book = openpyxl.load_workbook('coleta_nome_fantasia.xlsx') 
delorean_page = book['coleta_diaria']
bookAc = book.active



##############################


# Coleta de nome fantasia
def consulta_cnpj(cnpj):

    # configurações API
    url = f'https://brasilapi.com.br/api/cnpj/v1/{cnpj}'
    querystring = {"token":"XXXXXXXX-XXXX-XXXX-XXXX-XXXXXXXXXXXX","cnpj":"06990590000123","plugin":"RF"}
    response = requests.request("GET", url, params = querystring)
    resp = json.loads(response.text)
    nome_fantasia = resp.get('nome_fantasia', 'Chave não encontrada')

    if nome_fantasia == '':
        return '********'
    else:
        return nome_fantasia


##############################


#Pegar os valores de CNPJ
valores_coluna = []

for rows in delorean_page.iter_rows(min_row=2):
        
        valor = rows[0].value  # Conta por índice (começa com 0)
        valor = str(valor) #str --> possibilita uso do len

        if len(valor) == 14:
            valores_coluna.append(valor)
        if len(valor) == 16:
            valor = valor[:-2]
            valores_coluna.append(valor) #caso converta algun CNPJ para float


##############################


# Adicionar o nome fantasia dentro de uma lista
lista_nome_fantasia = []

for i in valores_coluna:
    i = str(i)

    nome_fantasia = consulta_cnpj(i)
    lista_nome_fantasia.append(nome_fantasia)


##############################


# Adicionar cada nome_fantasia dentro de uma célula
for i, valor in enumerate(lista_nome_fantasia, start=2):
    bookAc.cell(row=i, column=2, value=valor)

# Salvar o arquivo
book.save('RESULTADO_nome_fantasia.xlsx')
print("Valores adicionados com sucesso!")
    







